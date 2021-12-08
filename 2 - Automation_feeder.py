# importing necessary classes
# from different modules
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from zipfile import ZipFile
from pathlib import Path
import creds
import time
import os
import ctypes  # An included library with Python install.
import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter.messagebox import showerror

# root window
root = tk.Tk()
root.title('Batch Update')
root.geometry('300x130')
root.resizable(False, False)

frame = ttk.Frame(root)
options = {'padx': 5, 'pady': 5}

# startRow_label label
startRow_label = ttk.Label(frame, text='Start Row')
startRow_label.grid(column=0, row=0, sticky='W', **options)

# StartRow entry
startRowInput = tk.StringVar()
startRow_entry = ttk.Entry(frame, textvariable=startRowInput)
startRow_entry.grid(column=1, row=0, **options)
startRow_entry.focus()


# endRow_label label
startRow_label = ttk.Label(frame, text='End Row')
startRow_label.grid(column=0, row=1, sticky='W', **options)

endRowInput = tk.StringVar()
endRow_entry = ttk.Entry(frame, textvariable=endRowInput)
endRow_entry.grid(column=1, row=1, **options)
endRow_entry.focus()


def run_button_clicked():

    # open excel file
    currentPath = os.getcwd()
    path = currentPath + "//feedingsupload.xlsx"

    # open wb with data only for date format
    wb_obj = openpyxl.load_workbook(path, data_only=True)

    # open Tab
    sheet_obj = wb_obj.active

    # Getting the value of maximum rows
    # and column;
    rows = sheet_obj.max_row
    column = sheet_obj.max_column

    print("Total Rows:", rows)
    print("Total Columns:", column)

    # prompt user for start and end rows
    # startRowStr = input("Start Row: ")
    startRow = int(startRowInput.get())
    # endRowStr = input("End Row: ")
    endRow = int(endRowInput.get())

    chrome_options = webdriver.ChromeOptions()

    prefs = {"profile.default_content_setting_values.notifications": 2}
    chrome_options.add_experimental_option("prefs", prefs)
    browser = webdriver.Chrome("chromedriver.exe")

    print("Opening browser")

    # open reptilescan.com using get() method
    browser.get('https://login.reptilescan.com/')

    # get username and password
    username = creds.username
    password = creds.password

    print("Let's Begin")

    element = browser.find_elements_by_xpath('//*[@id="Input_Email"]')
    element[0].send_keys(username)

    print("Username Entered")

    element = browser.find_element_by_xpath('//*[@id="Input_Password"]')
    element.send_keys(password)

    print("Password Entered")

    # logging in
    # log_in = browser.find_elements_by_id('loginbutton')
    log_in = browser.find_elements_by_xpath(
        '//*[@id="LoginForm"]/form/div[3]/button')
    log_in[0].click()

    print("Login Successful")

    calc_records = endRow - startRow

    for i in range(startRow, endRow + 1):
        # cell_obj = sheet_obj.cell(row=2, column=i)
        # print(cell_obj.value, end=" ")

        status = 'Record #{i} of {calc_records}'
        result_label.config(text=status)

        id = str(sheet_obj.cell(row=i, column=1).value)
        print("id = " + id, end="\n ")

        idtracker = str(sheet_obj.cell(row=i, column=6).value)
        print("idtracker = " + idtracker, end="\n ")

        morph = str(sheet_obj.cell(row=i, column=5).value)
        print("morph = " + morph, end="\n ")

        # =TEXT(A2,mm/dd/yyyy")&" "&TEXT(A2,"hh:mm")
        date = str(sheet_obj.cell(row=i, column=9).value)
        print("date = " + date, end="\n ")

        name = str(sheet_obj.cell(row=i, column=3).value)
        print("name = " + name, end="\n ")

        size = str(sheet_obj.cell(row=i, column=4).value)
        print("size = " + size, end="\n ")

        count = str(sheet_obj.cell(row=i, column=2).value)
        print("count = " + count, end="\n ")

        # Open Event Page
        browser.get('https://login.reptilescan.com/Events/' +
                    id + '/New?category=feedings')

        # Selec the date, first clear it then enter the new one
        element = browser.find_elements_by_xpath('//*[@id="CreatedAt"]')
        element[0].clear()
        element[0].send_keys(date)

        # Select the name
        element = Select(browser.find_element_by_xpath(
            '//*[@id="EventTypeId"]'))
        element.select_by_visible_text(name)

        # Select the size
        element = Select(browser.find_element_by_xpath('//*[@id="Size"]'))
        element.select_by_visible_text(size)

        # Enter the count
        element = browser.find_elements_by_xpath('//*[@id="Count"]')
        element[0].send_keys(count)

        # Click the save button
        element = browser.find_element_by_xpath(
            '//*[@id="Content"]/main/form/div/div/div[2]/input')
        element.click()

        # Write to Excel sheet that it's been updated
        sheet_obj.cell(row=i, column=10).value = "Uploaded"

    # Save Spreadsheet
    wb_obj.save('reptilefeedings.xlsx')
    print("done")

    # Close the browser
    browser.close()


# Run button
run_button = ttk.Button(frame, text='Run')
run_button.grid(column=2, row=1, sticky='W', **options)
run_button.configure(command=run_button_clicked)

# result label
result_label = ttk.Label(frame)
result_label.grid(row=2, columnspan=3, **options)

frame.grid(padx=10, pady=10)

root.mainloop()
